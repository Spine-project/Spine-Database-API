######################################################################################################################
# Copyright (C) 2017-2021 Spine project consortium
# This file is part of Spine Database API.
# Spine Database API is free software: you can redistribute it and/or modify it under the terms of the GNU Lesser
# General Public License as published by the Free Software Foundation, either version 3 of the License, or (at your
# option) any later version. This program is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY;
# without even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU Lesser General
# Public License for more details. You should have received a copy of the GNU Lesser General Public License along with
# this program. If not, see <http://www.gnu.org/licenses/>.
######################################################################################################################

"""
A writer for exporting Spine databases to Excel files.

:author: A. Soininen (VTT)
:date:   15.1.2021
"""
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException
from .writer import Writer, WriterException


class ExcelWriter(Writer):
    def __init__(self, file_path):
        """
        Args:
            file_path (str): path ot output file
        """
        super().__init__()
        self._file_path = file_path
        self._workbook = None
        self._current_sheet = None
        self._removable_sheet_names = set()

    def finish(self):
        """See base class."""
        if self._workbook is None:
            return
        for name in self._removable_sheet_names:
            self._workbook.remove(self._workbook[name])
        self._workbook.save(self._file_path)
        self._workbook.close()
        self._workbook = None

    def finish_table(self):
        """See base class."""
        self._current_sheet = None

    def start(self):
        """See base class."""
        if Path(self._file_path).exists():
            try:
                self._workbook = load_workbook(self._file_path)
            except InvalidFileException as e:
                raise WriterException(f"Cannot open Excel file: {e}")
        else:
            self._workbook = Workbook()
            self._removable_sheet_names = set(self._workbook.sheetnames)

    def start_table(self, table_name):
        """See base class."""
        self._current_sheet = self._workbook.create_sheet(table_name)
        try:
            self._removable_sheet_names.remove(self._current_sheet.title)
        except KeyError:
            pass

    def write_row(self, row):
        """See base class."""
        self._current_sheet.append(row)
        return True